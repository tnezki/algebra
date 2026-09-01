\
#!/usr/bin/env python3

from __future__ import annotations

import html
import re
import tempfile
from datetime import date, datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

SPREADSHEET_ID = "1Qga2eTz0Nfgu8wIw5L-ZGC2xRtb4fOUoUXN-TyxuASE"
SHEET_NAME = "Student Calendar"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"

HERE = Path(__file__).resolve().parent
OUTPUT = HERE / "index.html"

RESOURCE_LINKS = [
    ("Course Website", "https://tnezki.github.io/algebra/"),
    ("Overview", "https://tnezki.github.io/algebra/misc/overview.html"),
    ("Textbook", "https://tnezki.github.io/textbooks/alg/index.html"),
    ("Virtual Tools", "https://technology.cpm.org/general/tiles/"),
    ("Printables", "https://tnezki.github.io/algebra/misc/printables/aaagallery_index.html"),
    ("Desmos", "https://www.desmos.com/calculator"),
    ("GeoGebra", "https://www.geogebra.org/graphing"),
    ("Canvas", "https://mariners.instructure.com/"),
    ("Upload Spot", "https://drive.google.com/drive/folders/1DwDKsvsAHMFefLMderdK8PI3MBcvXAxu?usp=sharing"),
]

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]


def safe_text(value):
    if value is None or value.__class__.__name__ == "ArrayFormula":
        return ""
    if isinstance(value, (datetime, date)):
        return f"{value.month}/{value.day}"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def looks_like_date(value):
    if isinstance(value, (datetime, date)):
        return True
    if value is None:
        return False
    return bool(re.fullmatch(r"\d{1,2}/\d{1,2}(?:/\d{2,4})?", str(value).strip()))


def date_key(value, school_start_year=2026):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?", str(value).strip())
    if not m:
        return None
    month, day = int(m.group(1)), int(m.group(2))
    year_text = m.group(3)
    if year_text:
        year = int(year_text)
        if year < 100:
            year += 2000
    else:
        year = school_start_year if month >= 7 else school_start_year + 1
    try:
        return date(year, month, day)
    except ValueError:
        return None



def direct_cell_link(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target

    formula = cell.value
    if isinstance(formula, str) and formula.startswith("="):
        match = re.search(r'HYPERLINK\(\s*"([^"]+)"', formula, re.I)
        if match:
            return match.group(1)

    return None


def normalized_key(value):
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalized_label(value):
    return re.sub(r"\s+", " ", safe_text(value)).strip().casefold()


def build_date_to_pacing_key(teacher_values):
    mapping = {}

    # Teacher Calendar historical/current week blocks use dates in D:H.
    # The corresponding Pacing day number is repeated in I:M on the
    # content rows directly below each date row.
    for row in range(1, min(teacher_values.max_row, 600) + 1):
        date_values = [
            teacher_values.cell(row=row, column=col).value
            for col in range(4, 9)
        ]

        if sum(1 for value in date_values if looks_like_date(value)) < 4:
            continue

        for day_index, date_value in enumerate(date_values):
            key_date = date_key(date_value)
            if not key_date:
                continue

            key_col = 9 + day_index
            pacing_key = None

            # The key is normally in row+1, and is repeated down the block.
            # Search a few rows so holidays/blank first rows do not matter.
            for source_row in range(
                row + 1,
                min(row + 12, teacher_values.max_row + 1),
            ):
                candidate = normalized_key(
                    teacher_values.cell(
                        row=source_row,
                        column=key_col,
                    ).value
                )
                if candidate not in (None, ""):
                    pacing_key = candidate
                    break

            if pacing_key not in (None, ""):
                mapping[key_date] = pacing_key

    return mapping


def build_pacing_link_lookup(pacing_values, pacing_formulas):
    lookup = {}
    direct_link_count = 0

    for row in range(1, min(pacing_values.max_row, 500) + 1):
        pacing_key = normalized_key(
            pacing_values.cell(row=row, column=1).value
        )
        if pacing_key in (None, ""):
            continue

        row_links = lookup.setdefault(pacing_key, {})

        for col in range(2, min(pacing_values.max_column, 26) + 1):
            label = safe_text(
                pacing_values.cell(row=row, column=col).value
            )
            if not label:
                continue

            link = (
                direct_cell_link(
                    pacing_formulas.cell(row=row, column=col)
                )
                or direct_cell_link(
                    pacing_values.cell(row=row, column=col)
                )
            )

            if link:
                direct_link_count += 1
                row_links.setdefault(normalized_label(label), link)

    return lookup, direct_link_count


def cell_info(
    ws_values,
    ws_links,
    row,
    col,
    day_date=None,
    date_to_pacing_key=None,
    pacing_links=None,
):
    vc = ws_values.cell(row=row, column=col)
    lc = ws_links.cell(row=row, column=col)

    # Keep evaluated blank formulas blank instead of displaying the formula.
    value = vc.value
    label = safe_text(value)

    link = direct_cell_link(vc) or direct_cell_link(lc)

    # Google XLSX export drops hyperlinks inherited through formulas on
    # Student Calendar. Restore them by date/day-number + exact displayed label.
    if (
        not link
        and label
        and day_date is not None
        and date_to_pacing_key is not None
        and pacing_links is not None
    ):
        key_date = date_key(day_date)
        pacing_key = (
            date_to_pacing_key.get(key_date)
            if key_date is not None
            else None
        )

        if pacing_key is not None:
            link = pacing_links.get(pacing_key, {}).get(
                normalized_label(label)
            )

    return label, link


def gather_items(
    ws_values,
    ws_links,
    start_row,
    end_row,
    col,
    day_date=None,
    date_to_pacing_key=None,
    pacing_links=None,
):
    items = []

    for row in range(start_row, end_row + 1):
        label, url = cell_info(
            ws_values,
            ws_links,
            row,
            col,
            day_date=day_date,
            date_to_pacing_key=date_to_pacing_key,
            pacing_links=pacing_links,
        )
        if label:
            items.append((label, url))

    return items


def download_workbook():
    response = requests.get(
        EXPORT_URL,
        timeout=45,
        headers={"User-Agent": "Mozilla/5.0 AgendaBuilder/1.0"},
    )
    response.raise_for_status()
    if len(response.content) < 1000:
        raise RuntimeError("Google returned an unexpectedly small workbook export.")
    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    f.write(response.content)
    f.close()
    return Path(f.name)



def read_calendar():
    xlsx = download_workbook()

    try:
        wb_values = load_workbook(
            xlsx,
            data_only=True,
            read_only=False,
        )
        wb_links = load_workbook(
            xlsx,
            data_only=False,
            read_only=False,
        )

        required = (SHEET_NAME, "Teacher Calendar", "Pacing")
        for sheet_name in required:
            if sheet_name not in wb_values.sheetnames:
                raise RuntimeError(f"Missing sheet: {sheet_name}")

        wsv = wb_values[SHEET_NAME]
        wsl = wb_links[SHEET_NAME]

        teacher_values = wb_values["Teacher Calendar"]
        pacing_values = wb_values["Pacing"]
        pacing_formulas = wb_links["Pacing"]

        date_to_pacing_key = build_date_to_pacing_key(
            teacher_values
        )
        pacing_links, direct_link_count = build_pacing_link_lookup(
            pacing_values,
            pacing_formulas,
        )

        print(
            f"Teacher dates mapped to Pacing days: "
            f"{len(date_to_pacing_key)}"
        )
        print(
            f"Pacing direct hyperlinks available: "
            f"{direct_link_count}"
        )

        if len(date_to_pacing_key) < 20:
            raise RuntimeError(
                "Too few Teacher Calendar dates mapped to Pacing."
            )

        if direct_link_count < 100:
            raise RuntimeError(
                "Too few direct Pacing hyperlinks were found."
            )

        current_date_values = [
            wsv.cell(row=2, column=col).value
            for col in range(1, 6)
        ]
        current_dates = [
            safe_text(value)
            for value in current_date_values
        ]

        current_items = [
            gather_items(
                wsv,
                wsl,
                3,
                9,
                col,
                day_date=current_date_values[col - 1],
                date_to_pacing_key=date_to_pacing_key,
                pacing_links=pacing_links,
            )
            for col in range(1, 6)
        ]

        current_start = next(
            (
                date_key(value)
                for value in current_date_values
                if date_key(value)
            ),
            None,
        )

        archive_date_rows = []

        for row in range(11, min(wsv.max_row, 500) + 1):
            values = [
                wsv.cell(row=row, column=col).value
                for col in range(1, 6)
            ]

            if sum(
                1 for value in values if looks_like_date(value)
            ) >= 4:
                archive_date_rows.append(row)

        previous = []

        for i, date_row in enumerate(archive_date_rows):
            next_date_row = (
                archive_date_rows[i + 1]
                if i + 1 < len(archive_date_rows)
                else min(date_row + 11, wsv.max_row + 1)
            )

            archive_date_values = [
                wsv.cell(row=date_row, column=col).value
                for col in range(1, 6)
            ]
            dates = [
                safe_text(value)
                for value in archive_date_values
            ]

            start_date = next(
                (
                    date_key(value)
                    for value in archive_date_values
                    if date_key(value)
                ),
                None,
            )

            is_current = bool(
                current_start
                and start_date
                and start_date == current_start
            )

            end_row = min(
                next_date_row - 1,
                date_row + 10,
            )

            items = [
                gather_items(
                    wsv,
                    wsl,
                    date_row + 1,
                    end_row,
                    col,
                    day_date=archive_date_values[col - 1],
                    date_to_pacing_key=date_to_pacing_key,
                    pacing_links=pacing_links,
                )
                for col in range(1, 6)
            ]

            if any(items):
                previous.append(
                    {
                        "dates": dates,
                        "items": items,
                        "is_current": is_current,
                    }
                )

        return {
            "current": {
                "dates": current_dates,
                "items": current_items,
            },
            "previous": previous,
        }

    finally:
        try:
            xlsx.unlink()
        except OSError:
            pass


def render_link(label, url, kind=""):
    classes = "cal-link" + (f" {kind}" if kind else "")
    if url:
        return (
            f'<a class="{classes}" href="{html.escape(url, quote=True)}" '
            f'target="_blank" rel="noopener">{html.escape(label)}</a>'
        )
    return f'<span class="{classes} no-link">{html.escape(label)}</span>'


def render_week(dates, items_by_day, current=False, archive_current=False):
    if current:
        cells = "".join(
            f"<th><div class='dow'>{d}</div><div class='date'>{html.escape(x)}</div></th>"
            for d, x in zip(DAY_NAMES, dates)
        )
        header = f'<tr class="week-head">{cells}</tr>'
        cls = "current-week"
    else:
        cells = "".join(
            f"<th><div class='date'>{html.escape(x)}</div></th>"
            for x in dates
        )
        header = f'<tr class="week-head">{cells}</tr>'
        cls = "calendar-current-week" if archive_current else "previous-week"

    row_count = max([len(x) for x in items_by_day] + [1])
    body_rows = []

    for i in range(row_count):
        tds = []
        for items in items_by_day:
            if i >= len(items):
                tds.append("<td></td>")
                continue

            label, url = items[i]
            kind = ""
            if i == 0:
                low = label.lower()
                kind = "holiday" if any(x in low for x in ("labor day", "pd", "no school", "holiday")) else "lesson"
            tds.append(f"<td>{render_link(label, url, kind)}</td>")
        body_rows.append("<tr>" + "".join(tds) + "</tr>")

    return f'<tbody class="week-block {cls}">{header}{"".join(body_rows)}</tbody>'


def build_html(calendar):
    resources = "\n".join(
        f'<a href="{html.escape(url, quote=True)}" target="_blank" rel="noopener">{html.escape(label)}</a>'
        for label, url in RESOURCE_LINKS
    )

    current_html = render_week(
        calendar["current"]["dates"],
        calendar["current"]["items"],
        current=True,
    )

    previous_html = ""
    if calendar["previous"]:
        previous_html = (
            '<tbody class="previous-weeks-divider"><tr>'
            '<td colspan="5">Previous Weeks</td>'
            '</tr></tbody>'
            + "".join(
                render_week(
                    w["dates"],
                    w["items"],
                    current=False,
                    archive_current=w.get("is_current", False),
                )
                for w in calendar["previous"]
            )
        )

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta http-equiv="refresh" content="300">
<title>Algebra 1 Agenda 2026-2027</title>
<style>
:root {{
  --navy:#173f6d;
  --navy-dark:#173f6d;
  --gold:#e0bd4f;
  --gold-light:#fff0b8;
  --gold-pale:#fff8df;
  --ink:#1f2937;
  --muted:#64748b;
  --lesson:#fff0b8;
  --link:#173f6d;
}}
* {{ box-sizing:border-box; }}
body {{
  margin:0;
  font-family:Arial, Helvetica, sans-serif;
  background:#fff;
  color:var(--ink);
}}
.wrapper {{
  width:min(1180px, calc(100% - 24px));
  margin:18px auto 40px;
}}
.titlebar {{
  background:var(--navy);
  color:#fff;
  padding:16px 22px;
  border-radius:10px 10px 0 0;
  text-align:center;
}}
.titlebar h1 {{
  margin:0;
  font-size:clamp(1.55rem, 3vw, 2.25rem);
  font-weight:800;
}}
.titlebar .small {{
  font-size:.64em;
  font-weight:600;
}}
.resources {{
  border:1px solid var(--gold);
  border-top:0;
  padding:12px 14px 13px;
  display:flex;
  flex-wrap:wrap;
  justify-content:center;
  gap:8px;
  background:#fff9e9;
}}
.resources a {{
  text-decoration:none;
  color:var(--navy-dark);
  background:#fff;
  border:1px solid var(--gold);
  border-radius:999px;
  padding:7px 11px;
  font-size:.88rem;
  font-weight:700;
}}
.resources a:hover,
.resources a:focus-visible {{
  background:var(--gold-light);
  border-color:var(--navy);
}}
.calendar-wrap {{
  overflow-x:auto;
  border:1px solid #c8b675;
  border-top:0;
}}
table {{
  border-collapse:collapse;
  width:100%;
  min-width:760px;
  table-layout:fixed;
}}
th, td {{
  border-right:1px solid #cfd4da;
  border-bottom:1px solid #cfd4da;
  text-align:center;
  vertical-align:middle;
}}
tr > *:last-child {{ border-right:0; }}
.week-head th {{ padding:8px 6px; }}
.dow {{ font-size:.9rem; font-weight:800; }}
.date {{ margin-top:2px; font-size:.82rem; font-weight:700; }}
td {{ padding:5px 6px; background:#fff; height:34px; }}
.cal-link {{
  display:block;
  width:100%;
  text-decoration:none;
  color:var(--link);
  font-size:.88rem;
  font-weight:650;
  padding:4px 5px;
  border-radius:5px;
}}
a.cal-link:hover,
a.cal-link:focus-visible {{
  background:#fff4c7;
  text-decoration:none;
}}
.cal-link.lesson {{
  background:var(--lesson);
  border:1px solid #e1c86d;
  font-weight:800;
  color:var(--navy-dark);
}}
.cal-link.holiday {{
  background:#f6edcf;
  color:#475569;
  font-weight:800;
}}
.no-link {{ cursor:default; }}

.current-week .week-head th {{
  background:var(--navy);
  color:#fff;
  text-align:left;
  padding:9px 10px 10px;
}}
.current-week .dow {{
  font-size:1.02rem;
  font-weight:850;
  line-height:1.1;
}}
.current-week .date {{
  margin-top:4px;
  font-size:1.22rem;
  font-weight:900;
  text-align:left;
}}
.current-week td {{
  padding:0;
  height:48px;
  min-height:48px;
  background:#fff;
}}
.current-week .cal-link {{
  display:flex;
  align-items:center;
  justify-content:center;
  width:100%;
  min-height:48px;
  padding:10px 8px;
  font-size:1.08rem;
  font-weight:750;
  line-height:1.25;
}}
.current-week .cal-link.lesson {{
  background:#fff;
  border:0;
  color:var(--navy);
  font-size:1.12rem;
  font-weight:900;
  text-decoration:none;
}}
.current-week .cal-link.holiday {{
  font-size:1.08rem;
  font-weight:900;
}}
.current-week tr:nth-child(odd):not(.week-head) td {{
  background:var(--gold-light);
}}

.previous-weeks-divider td {{
  background:var(--gold) !important;
  color:#10243d;
  font-size:1.05rem;
  font-weight:800;
  padding:10px 8px;
}}

.calendar-current-week .week-head th {{
  background:var(--navy);
  color:#fff;
  border-top:5px solid var(--gold);
  padding:8px 9px;
  font-weight:850;
  text-align:left;
}}
.calendar-current-week .week-head th .date {{
  color:#fff;
  font-size:1rem;
  font-weight:900;
  text-align:left;
}}
.calendar-current-week tr td {{
  background:#fff !important;
  border-color:#cfd4da;
}}
.calendar-current-week tr:nth-child(odd):not(.week-head) td {{
  background:var(--gold-light) !important;
}}
.calendar-current-week .cal-link.lesson {{
  background:#fff;
  border:0;
  color:var(--navy);
  font-weight:900;
  text-decoration:none;
}}
.calendar-current-week .cal-link.holiday {{
  background:#f6edcf;
  color:#475569;
}}

.previous-week .week-head th {{
  background:#3f4650;
  color:#fff;
  border-top:5px solid #20242a;
  text-align:left;
  padding-left:9px;
}}
.previous-week .week-head th .date {{ color:#e5e7eb; }}
.previous-week tr td {{
  background:#fff !important;
  border-color:#cfd4da;
}}
.previous-week tr:nth-child(even) td {{
  background:#f3f4f6 !important;
}}
.previous-week .cal-link.lesson {{
  background:#e5e7eb;
  border:1px solid #c7ccd1;
  color:#2f3740;
}}
.previous-week .cal-link.holiday {{
  background:#eceff2;
  color:#4b5563;
}}
.previous-week a.cal-link:hover,
.previous-week a.cal-link:focus-visible {{
  background:#e2e6ea;
}}
.updated {{
  text-align:center;
  color:var(--muted);
  font-size:.76rem;
  padding-top:8px;
}}

@media (max-width:700px) {{
  .wrapper {{ width:100%; margin:0; }}
  .titlebar {{ border-radius:0; }}
  .resources {{ justify-content:flex-start; padding:10px; }}
  .resources a {{ font-size:.8rem; padding:6px 9px; }}
}}
</style>
</head>
<body>
<div class="wrapper">
  <header class="titlebar">
    <h1>Algebra 1 <span class="small">– Agenda 2026-2027</span></h1>
  </header>

  <nav class="resources" aria-label="Student resources">
    {resources}
  </nav>

  <div class="calendar-wrap">
    <table aria-label="Algebra 1 student agenda">
      {current_html}
      {previous_html}
    </table>
  </div>

  <div class="updated">Agenda generated {html.escape(stamp)}</div>
</div>
</body>
</html>
"""


def main():
    calendar = read_calendar()
    output = build_html(calendar)
    temp = OUTPUT.with_suffix(".html.tmp")
    temp.write_text(output, encoding="utf-8")
    temp.replace(OUTPUT)
    print(f"Updated {OUTPUT}")


if __name__ == "__main__":
    main()
